"""Excel 报告生成模块。

提供 generate_xlsx 函数，将测量数据和分析结果写入格式化的
Excel 汇总表格。
"""

from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Union

import numpy as np

from echem_core.model import Measurement, Technique
from echem_core.analysis.lsv import find_e_half, tafel_slope
from echem_core.analysis.cv import calc_cdl, calc_ecsa


def _get_plot_data(
    measurement: Measurement,
) -> tuple:
    """获取用于分析的电位/电流数据（优先使用处理后的数据）。"""
    potential = (
        measurement.processed_potential
        if measurement.processed_potential is not None
        else measurement.raw_potential
    )
    current = (
        measurement.processed_current
        if measurement.processed_current is not None
        else measurement.raw_current
    )
    return potential, current


def _safe_analysis(callable, *args, **kwargs):
    """安全执行分析函数，失败时返回 'N/A'。"""
    try:
        result = callable(*args, **kwargs)
        return result
    except Exception:
        return "N/A"


def _compute_file_hash(measurement: Measurement) -> str:
    """从 Measurement 对象获取或计算文件哈希。"""
    if measurement.file_hash is not None:
        return measurement.file_hash
    return "N/A"


def _format_e_half(result) -> str:
    """格式化半波电位分析结果。"""
    if result == "N/A":
        return "N/A"
    e_half, j_L, confidence = result
    return f"{e_half:.4f}"


def _format_j_L(result) -> str:
    """格式化极限电流结果。"""
    if result == "N/A":
        return "N/A"
    e_half, j_L, confidence = result
    return f"{j_L:.4f}"


def _format_tafel(result) -> str:
    """格式化 Tafel 斜率结果。"""
    if result == "N/A":
        return "N/A"
    slope, intercept, r_squared, region_start, region_end = result
    return f"{slope:.3f}"


def _format_cdl(result) -> str:
    """格式化 Cdl 结果。"""
    if result == "N/A":
        return "N/A"
    cdl, r_squared, df_dv_values, scan_rates = result
    return f"{cdl:.6f}"


def _format_ecsa(cdl_result, specific_capacitance: float = 0.04) -> str:
    """从 Cdl 结果计算并格式化 ECSA。"""
    if cdl_result == "N/A":
        return "N/A"
    cdl, r_squared, df_dv_values, scan_rates = cdl_result
    try:
        ecsa = calc_ecsa(cdl, specific_capacitance=specific_capacitance)
        return f"{ecsa:.4f}"
    except Exception:
        return "N/A"


def generate_xlsx(
    measurements: Sequence[Measurement],
    path: Union[str, Path],
    specific_capacitance: float = 0.04,
) -> Path:
    """生成 Excel 分析汇总报告。

    为每组测量数据计算关键电化学参数并写入格式化表格。

    Excel 包含列:
        sample_name, technique, E1/2, j_L, Tafel_slope, Cdl, ECSA, file_hash

    Args:
        measurements: Measurement 对象序列。
        path: 输出 .xlsx 文件路径（父目录不存在会自动创建）。
        specific_capacitance: ECSA 计算用的比电容值（mF/cm²），
            默认 0.04（适用于碱性环境金属表面）。

    Returns:
        生成的 Excel 文件路径。

    Raises:
        ValueError: measurements 为空时抛出。
        ImportError: openpyxl 未安装时抛出。
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise ImportError(
            "需要 openpyxl 库来生成 Excel 报告。"
            "请运行: pip install openpyxl"
        )

    measurements = list(measurements)
    if not measurements:
        raise ValueError("measurements 列表为空，无法生成报告")

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # ── 创建工作簿 ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分析汇总"

    # ── 表头样式 ────────────────────────────────────────────────────────────
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── 写表头 ──────────────────────────────────────────────────────────────
    headers = [
        "sample_name",
        "technique",
        "E1/2 (V)",
        "j_L (mA/cm²)",
        "Tafel_slope (mV/dec)",
        "Cdl (F)",
        "ECSA (cm²)",
        "file_hash",
    ]
    col_widths = [30, 12, 14, 16, 22, 14, 14, 72]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # ── 冻结首行 ────────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── 数据行样式 ──────────────────────────────────────────────────────────
    data_font = Font(name="Consolas", size=10)
    data_alignment = Alignment(horizontal="center", vertical="center")
    na_font = Font(name="Consolas", size=10, color="999999", italic=True)

    # ── 写入数据 ────────────────────────────────────────────────────────────
    for row_idx, m in enumerate(measurements, start=2):
        potential, current = _get_plot_data(m)
        sample_name = m.metadata.get("sample_name", "") or ""
        technique = m.technique.value
        file_hash = _compute_file_hash(m)

        # --- LSV 分析: E1/2, j_L, Tafel_slope ---
        if m.technique in (Technique.LSV,):
            e_half_result = _safe_analysis(find_e_half, potential, current)
            e_half_str = _format_e_half(e_half_result)
            j_L_str = _format_j_L(e_half_result)

            # Tafel 斜率（tafel_slope 已返回 mV/dec）
            j_L_val = e_half_result[1] if isinstance(e_half_result, tuple) else None
            if j_L_val is not None and j_L_val != 0:
                tafel_result = _safe_analysis(tafel_slope, potential, current, j_L_val)
                tafel_str = _format_tafel(tafel_result)
            else:
                tafel_str = "N/A"

        else:
            e_half_str = "N/A"
            j_L_str = "N/A"
            tafel_str = "N/A"

        # --- CV 分析: Cdl, ECSA ---
        if m.technique in (Technique.CV,):
            # Cdl 需要多个不同扫速的测量，这里只传当前一个，
            # 所以需要外部先计算 Cdl 再传入。
            # 此处写 "N/A" 并提示用户。
            cdl_str = "N/A"
            ecsa_str = "N/A"
        else:
            cdl_str = "N/A"
            ecsa_str = "N/A"

        # ── 写入单元格 ──────────────────────────────────────────────
        values = [
            sample_name,
            technique,
            e_half_str,
            j_L_str,
            tafel_str,
            cdl_str,
            ecsa_str,
            file_hash,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            is_na = (isinstance(value, str) and value == "N/A")
            cell.font = na_font if is_na else data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    # ── 自动筛选 ────────────────────────────────────────────────────────────
    last_col_letter = openpyxl.utils.get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(measurements) + 1}"

    # ── 保存 ────────────────────────────────────────────────────────────────
    wb.save(str(path))
    print(f"[Report] Excel 报告已保存: {path}")
    return path


def generate_xlsx_with_cdl(
    measurements: Sequence[Measurement],
    cv_measurements: Sequence[Measurement],
    path: Union[str, Path],
    specific_capacitance: float = 0.04,
) -> Path:
    """生成包含 Cdl / ECSA 分析的完整 Excel 报告。

    CV 测量需要多个不同扫描速率的数据，通过 calc_cdl() 计算
    双电层电容后，再将 ECSA 写入所有 LSV 测量对应的行。

    Args:
        measurements: 所有要报告的 Measurement 对象。
        cv_measurements: 用于计算 Cdl 的 CV 测量列表（不同扫速）。
        path: 输出 .xlsx 文件路径。
        specific_capacitance: ECSA 计算比电容（mF/cm²），默认 0.04。

    Returns:
        生成的 Excel 文件路径。
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise ImportError("需要 openpyxl 库来生成 Excel 报告。")

    measurements = list(measurements)
    if not measurements:
        raise ValueError("measurements 列表为空")

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # 计算 Cdl
    cdl_result = _safe_analysis(calc_cdl, list(cv_measurements))
    cdl_str = _format_cdl(cdl_result)
    ecsa_str = _format_ecsa(cdl_result, specific_capacitance=specific_capacitance)

    # ── 创建工作簿 ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分析汇总"

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "sample_name",
        "technique",
        "E1/2 (V)",
        "j_L (mA/cm²)",
        "Tafel_slope (mV/dec)",
        "Cdl (F)",
        "ECSA (cm²)",
        "file_hash",
    ]
    col_widths = [30, 12, 14, 16, 22, 14, 14, 72]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    data_font = Font(name="Consolas", size=10)
    data_alignment = Alignment(horizontal="center", vertical="center")
    na_font = Font(name="Consolas", size=10, color="999999", italic=True)

    for row_idx, m in enumerate(measurements, start=2):
        potential, current = _get_plot_data(m)
        sample_name = m.metadata.get("sample_name", "") or ""
        technique = m.technique.value
        file_hash = _compute_file_hash(m)

        # LSV 分析
        if m.technique in (Technique.LSV,):
            e_half_result = _safe_analysis(find_e_half, potential, current)
            e_half_str = _format_e_half(e_half_result)
            j_L_str = _format_j_L(e_half_result)
            j_L_val = e_half_result[1] if isinstance(e_half_result, tuple) else None
            if j_L_val is not None and j_L_val != 0:
                tafel_result = _safe_analysis(tafel_slope, potential, current, j_L_val)
                tafel_str = _format_tafel(tafel_result)
            else:
                tafel_str = "N/A"
        else:
            e_half_str = "N/A"
            j_L_str = "N/A"
            tafel_str = "N/A"

        values = [
            sample_name,
            technique,
            e_half_str,
            j_L_str,
            tafel_str,
            cdl_str,
            ecsa_str,
            file_hash,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            is_na = (isinstance(value, str) and value == "N/A")
            cell.font = na_font if is_na else data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    last_col_letter = openpyxl.utils.get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(measurements) + 1}"

    wb.save(str(path))
    print(f"[Report] Excel 报告（含 Cdl/ECSA）已保存: {path}")
    return path


__all__ = ["generate_xlsx", "generate_xlsx_with_cdl"]
